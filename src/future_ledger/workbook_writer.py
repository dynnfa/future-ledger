"""Write FutureLedger report tables to the v0 Excel workbook."""

from __future__ import annotations

from collections.abc import Iterable, Mapping, Sequence
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Font  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]
from openpyxl.writer.excel import ExcelWriter  # type: ignore[import-untyped]

from future_ledger.domain import (
    DividendLongRow,
    DividendRankRow,
    MetadataRow,
    PricePoint,
    ReportTables,
    SourceErrorRow,
)
from future_ledger.errors import ConfigError

DIVIDEND_RANK_CORE_COLUMNS = (
    "rank_latest_yield",
    "stock_code",
    "stock_name",
    "market",
    "latest_report_year",
    "latest_cash_dividend_per_10_shares",
    "latest_cash_dividend_per_share",
    "reference_price",
    "reference_price_date",
    "latest_dividend_yield_pct",
    "dividend_yield_source",
    "dividend_year_count_5y",
    "continuous_dividend_5y",
    "avg_dividend_yield_pct_5y",
    "min_dividend_yield_pct_5y",
    "max_dividend_yield_pct_5y",
    "as_of_date",
    "cash_dividends_1y",
    "total_return_1y_pct",
    "annualized_return_1y_pct",
    "has_missing_years_5y",
    "data_quality_flags",
    "source_priority_used",
    "fetched_at",
)

DIVIDEND_LONG_COLUMNS = (
    "stock_code",
    "stock_name",
    "market",
    "report_year",
    "report_period",
    "cash_dividend_per_10_shares",
    "cash_dividend_per_share",
    "ex_dividend_date",
    "registration_date",
    "plan_status",
    "eps",
    "net_asset_per_share",
    "profit_growth_yoy_pct",
    "dividend_yield_pct",
    "source",
)

PRICE_POINT_COLUMNS = (
    "stock_code",
    "date",
    "close",
)

SOURCE_ERROR_COLUMNS = (
    "stock_code",
    "stage",
    "message",
    "raw_detail",
)

METADATA_COLUMNS = (
    "key",
    "value",
)

DATE_FORMAT = "yyyy-mm-dd"
PERCENT_UNIT_FORMAT = '0.00"%"'
DECIMAL_FORMAT = "0.0000"

TypedRow = DividendLongRow | PricePoint | SourceErrorRow | MetadataRow


def write_workbook(
    tables: ReportTables,
    output_path: str | Path,
    *,
    workbook_timestamp: datetime | None = None,
) -> Path:
    """Write report tables to an Excel workbook and return the written path."""
    path = Path(output_path)
    _validate_output_path(path)

    workbook = Workbook()
    workbook.remove(workbook.active)

    if workbook_timestamp is not None:
        workbook.properties.created = workbook_timestamp
        workbook.properties.modified = workbook_timestamp

    rank_columns = _rank_columns(tables.dividend_rank)
    _write_sheet(workbook, "dividend_rank", rank_columns, _rank_rows(tables.dividend_rank))
    _write_sheet(
        workbook,
        "dividend_long",
        DIVIDEND_LONG_COLUMNS,
        _mapped_rows(tables.dividend_long, DIVIDEND_LONG_COLUMNS),
    )
    _write_sheet(
        workbook,
        "price_points",
        PRICE_POINT_COLUMNS,
        _mapped_rows(tables.price_points, PRICE_POINT_COLUMNS),
    )
    _write_sheet(
        workbook,
        "source_errors",
        SOURCE_ERROR_COLUMNS,
        _mapped_rows(tables.source_errors, SOURCE_ERROR_COLUMNS),
    )
    _write_sheet(
        workbook,
        "metadata",
        METADATA_COLUMNS,
        _mapped_rows(tables.metadata, METADATA_COLUMNS),
    )

    try:
        path.parent.mkdir(parents=True, exist_ok=True)
        _save_workbook(workbook, path)
    except OSError as exc:
        raise ConfigError("failed to write workbook") from exc

    return path


def _save_workbook(workbook: Any, path: Path) -> None:
    archive = ZipFile(path, "w", ZIP_DEFLATED, allowZip64=True)
    writer = ExcelWriter(workbook, archive)
    writer.save()


def _validate_output_path(path: Path) -> None:
    if path.suffix != ".xlsx":
        raise ConfigError("--output must end with .xlsx")
    if path.parent.exists() and not path.parent.is_dir():
        raise ConfigError("output parent is not a directory")


def _rank_columns(rows: Sequence[DividendRankRow]) -> tuple[str, ...]:
    annual_columns: list[str] = []
    seen = set(DIVIDEND_RANK_CORE_COLUMNS)
    for row in rows:
        for column in row.annual_fields:
            if column not in seen:
                annual_columns.append(column)
                seen.add(column)
    return (*DIVIDEND_RANK_CORE_COLUMNS, *annual_columns)


def _rank_rows(rows: Sequence[DividendRankRow]) -> Iterable[Mapping[str, object]]:
    for row in rows:
        data = _row_mapping(row, DIVIDEND_RANK_CORE_COLUMNS)
        data.update(row.annual_fields)
        yield data


def _mapped_rows(
    rows: Sequence[TypedRow],
    columns: Sequence[str],
) -> Iterable[Mapping[str, object]]:
    for row in rows:
        yield _row_mapping(row, columns)


def _row_mapping(row: object, columns: Sequence[str]) -> dict[str, object]:
    return {column: getattr(row, column) for column in columns}


def _write_sheet(
    workbook: Any,
    title: str,
    columns: Sequence[str],
    rows: Iterable[Mapping[str, object]],
) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.freeze_panes = "A2"
    worksheet.append(list(columns))

    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        worksheet.append([_excel_value(row.get(column)) for column in columns])

    _format_sheet(worksheet, columns)


def _excel_value(value: object) -> object:
    if value is None:
        return None
    if isinstance(value, tuple):
        return "|".join(str(item) for item in value)
    if isinstance(value, Decimal):
        return value
    if isinstance(value, date):
        return value
    return value


def _format_sheet(worksheet: Any, columns: Sequence[str]) -> None:
    for index, column_name in enumerate(columns, start=1):
        letter = get_column_letter(index)
        worksheet.column_dimensions[letter].width = _column_width(column_name)
        if worksheet.max_row < 2:
            continue

        for cell in worksheet.iter_cols(
            min_col=index,
            max_col=index,
            min_row=2,
            max_row=worksheet.max_row,
        ):
            for item in cell:
                if item.value is None:
                    continue
                if _is_date_column(column_name):
                    item.number_format = DATE_FORMAT
                elif _is_percent_unit_column(column_name):
                    item.number_format = PERCENT_UNIT_FORMAT
                elif isinstance(item.value, Decimal):
                    item.number_format = DECIMAL_FORMAT


def _column_width(column_name: str) -> int:
    return min(max(len(column_name) + 2, 12), 32)


def _is_date_column(column_name: str) -> bool:
    return column_name == "date" or column_name.endswith("_date")


def _is_percent_unit_column(column_name: str) -> bool:
    return column_name.endswith("_pct")
