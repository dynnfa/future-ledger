from __future__ import annotations

import inspect
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook  # type: ignore[import-untyped]

from future_ledger.domain import (
    DividendLongRow,
    DividendRankRow,
    MetadataRow,
    PricePoint,
    ReportTables,
    SourceErrorRow,
)
from future_ledger.errors import ConfigError
from future_ledger.workbook_writer import write_workbook


def test_write_workbook_creates_required_sheets_in_order(tmp_path: Path) -> None:
    output = write_workbook(_tables(), tmp_path / "report.xlsx")

    workbook = load_workbook(output)

    assert workbook.sheetnames == [
        "dividend_rank",
        "dividend_long",
        "price_points",
        "source_errors",
        "metadata",
    ]


def test_write_workbook_preserves_rank_column_order(tmp_path: Path) -> None:
    output = write_workbook(_tables(), tmp_path / "report.xlsx")

    workbook = load_workbook(output)
    headers = [cell.value for cell in workbook["dividend_rank"][1]]

    assert headers == [
        "rank_latest_yield",
        "stock_code",
        "stock_name",
        "market",
        "latest_report_year",
        "latest_cash_dividend_per_10_shares",
        "latest_cash_dividend_per_share",
        "reference_price",
        "reference_price_date",
        "latest_dividend_yield_pct",
        "dividend_yield_source",
        "dividend_year_count_5y",
        "continuous_dividend_5y",
        "avg_dividend_yield_pct_5y",
        "min_dividend_yield_pct_5y",
        "max_dividend_yield_pct_5y",
        "as_of_date",
        "cash_dividends_1y",
        "total_return_1y_pct",
        "annualized_return_1y_pct",
        "has_missing_years_5y",
        "data_quality_flags",
        "source_priority_used",
        "fetched_at",
        "2025_report_period",
        "2025_cash_dividend_per_10_shares",
        "2025_reference_price_date",
    ]


def test_write_workbook_serializes_domain_values(tmp_path: Path) -> None:
    output = write_workbook(_tables(), tmp_path / "report.xlsx")

    workbook = load_workbook(output)
    rank = workbook["dividend_rank"]
    long = workbook["dividend_long"]
    prices = workbook["price_points"]
    errors = workbook["source_errors"]

    assert rank["A2"].value == 1
    assert rank["J2"].value == 5.2
    assert rank["Q2"].value == datetime(2026, 4, 20)
    assert rank["V2"].value == "missing_reference_price|missing_return_price"
    assert rank["AA2"].value == datetime(2025, 7, 1)
    assert long["N2"].value == 5.2
    assert prices["B2"].value == datetime(2025, 7, 1)
    assert prices["C2"].value == 10.25
    assert errors["D2"].value is None


def test_write_workbook_formats_percent_unit_values_without_scaling(tmp_path: Path) -> None:
    output = write_workbook(_tables(), tmp_path / "report.xlsx")

    workbook = load_workbook(output)
    rank = workbook["dividend_rank"]

    assert rank["J2"].value == 5.2
    assert rank["J2"].number_format == '0.00"%"'
    assert rank["S2"].value == 6.5
    assert rank["S2"].number_format == '0.00"%"'


def test_write_workbook_uses_supplied_timestamp_for_document_properties(
    tmp_path: Path,
) -> None:
    timestamp = datetime(2026, 5, 14, 8, 30, 0)

    output = write_workbook(
        _tables(),
        tmp_path / "report.xlsx",
        workbook_timestamp=timestamp,
    )

    workbook = load_workbook(output)
    assert workbook.properties.created == timestamp
    assert workbook.properties.modified == timestamp


def test_write_workbook_writes_empty_tables_with_headers(tmp_path: Path) -> None:
    output = write_workbook(ReportTables.empty(), tmp_path / "empty.xlsx")

    workbook = load_workbook(output)

    assert workbook.sheetnames == [
        "dividend_rank",
        "dividend_long",
        "price_points",
        "source_errors",
        "metadata",
    ]
    assert workbook["dividend_rank"].max_row == 1
    assert workbook["dividend_rank"].max_column == 24
    assert workbook["dividend_long"].max_row == 1
    assert workbook["price_points"].max_row == 1
    assert workbook["source_errors"].max_row == 1
    assert workbook["metadata"].max_row == 1


def test_write_workbook_rejects_non_xlsx_output(tmp_path: Path) -> None:
    with pytest.raises(ConfigError, match=r"--output must end with \.xlsx"):
        write_workbook(ReportTables.empty(), tmp_path / "report.csv")


def test_write_workbook_creates_parent_directories(tmp_path: Path) -> None:
    output = write_workbook(ReportTables.empty(), tmp_path / "nested" / "dir" / "report.xlsx")

    assert output == tmp_path / "nested" / "dir" / "report.xlsx"
    assert output.exists()


def test_write_workbook_rejects_parent_path_that_is_file(tmp_path: Path) -> None:
    parent_file = tmp_path / "not-a-directory"
    parent_file.write_text("file", encoding="utf-8")

    with pytest.raises(ConfigError, match="output parent is not a directory"):
        write_workbook(ReportTables.empty(), parent_file / "report.xlsx")


def test_workbook_writer_does_not_import_source_or_metric_modules() -> None:
    import future_ledger.workbook_writer as workbook_writer

    source = inspect.getsource(workbook_writer)
    forbidden_fragments = (
        "future_ledger.sources",
        "future_ledger.metrics",
        "future_ledger.normalize",
        "akshare",
        "pandas",
    )

    assert not any(fragment in source for fragment in forbidden_fragments)


def _tables() -> ReportTables:
    return ReportTables(
        dividend_rank=[
            DividendRankRow(
                rank_latest_yield=1,
                stock_code="600000",
                stock_name="Pudong Bank",
                market="SH",
                latest_report_year=2025,
                latest_cash_dividend_per_10_shares=Decimal("4.10"),
                latest_cash_dividend_per_share=Decimal("0.41"),
                reference_price=Decimal("10.25"),
                reference_price_date=date(2025, 7, 1),
                latest_dividend_yield_pct=Decimal("5.2"),
                dividend_yield_source="calculated_ex_dividend_close",
                dividend_year_count_5y=5,
                continuous_dividend_5y=True,
                avg_dividend_yield_pct_5y=Decimal("4.80"),
                min_dividend_yield_pct_5y=Decimal("3.10"),
                max_dividend_yield_pct_5y=Decimal("5.20"),
                as_of_date=date(2026, 4, 20),
                cash_dividends_1y=Decimal("0.41"),
                total_return_1y_pct=Decimal("6.5"),
                annualized_return_1y_pct=Decimal("6.5"),
                has_missing_years_5y=False,
                data_quality_flags=("missing_reference_price", "missing_return_price"),
                source_priority_used="akshare.stock_fhps_detail_em",
                fetched_at="2026-04-20T08:30:00+08:00",
                annual_fields={
                    "2025_report_period": "2025 annual",
                    "2025_cash_dividend_per_10_shares": Decimal("4.10"),
                    "2025_reference_price_date": date(2025, 7, 1),
                },
            )
        ],
        dividend_long=[
            DividendLongRow(
                stock_code="600000",
                stock_name="Pudong Bank",
                market="SH",
                report_year=2025,
                report_period="2025 annual",
                cash_dividend_per_10_shares=Decimal("4.10"),
                cash_dividend_per_share=Decimal("0.41"),
                ex_dividend_date=date(2025, 7, 1),
                registration_date=date(2025, 6, 30),
                plan_status="implemented",
                eps=Decimal("2.10"),
                net_asset_per_share=Decimal("18.50"),
                profit_growth_yoy_pct=Decimal("3.25"),
                dividend_yield_pct=Decimal("5.2"),
                source="akshare.stock_fhps_detail_em",
            )
        ],
        price_points=[
            PricePoint(
                stock_code="600000",
                date=date(2025, 7, 1),
                close=Decimal("10.25"),
            )
        ],
        source_errors=[
            SourceErrorRow(
                stock_code="600000",
                stage="price_fetch",
                message="empty upstream frame",
                raw_detail=None,
            )
        ],
        metadata=[
            MetadataRow(key="run.as_of", value="2026-04-20"),
            MetadataRow(
                key="disclaimer",
                value=(
                    "FutureLedger is for financial data research only. It does not "
                    "provide investment advice, buy/sell recommendations, or "
                    "portfolio allocation guidance."
                ),
            ),
        ],
    )
